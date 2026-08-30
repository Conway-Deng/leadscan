"""
report.py
---------
Turns the scored rows into the three things a caller actually uses:

  * warm_leads.csv   -- the machine-readable record
  * warm_leads.html  -- the call sheet. Open it in a browser and start calling.
  * warm_leads.xlsx  -- the same sheet for anyone who prefers a spreadsheet.

SPREADSHEET SAFETY
A business name can start with "=", "+", "-" or "@". Excel and Google Sheets
read such a value as a formula, which is a known way to attack a person who
opens a downloaded file. Every text value written to CSV or XLSX is therefore
given a leading apostrophe when it starts with one of those characters.
"""

import csv
import hashlib
import html as html_module
import os

import scoring

CSV_FIELDS = [
    "tier", "score", "name", "phone", "email", "review_count", "rating",
    "instagram_followers", "email_grade", "ad_tags", "capture_methods",
    "pages_checked", "hook",
    "website", "instagram", "facebook", "tiktok", "address", "opening_hours",
    "reasons", "status", "whatsapp_message", "email_subject", "email_body",
]

# Column heading, source key, and whether the value is a link.
SHEET_COLUMNS = [
    ("Tier", "tier", False),
    ("Score", "score", False),
    ("Business", "name", False),
    ("Phone", "phone", False),
    ("What to say", "hook", False),
    ("Advertising tags installed", "ad_tags", False),
    ("Reviews", "review_count", False),
    ("IG followers", "instagram_followers", False),
    ("Email", "email", False),
    ("Email quality", "email_grade", False),
    ("Opening hours", "opening_hours", False),
    ("Website", "website", True),
    ("Instagram", "instagram", True),
    ("Facebook", "facebook", True),
    ("TikTok", "tiktok", True),
    ("Status", "status", False),
]

_DANGEROUS_PREFIX = ("=", "+", "-", "@", "\t", "\r")


def safe_cell(value):
    """Stop a spreadsheet from reading a text value as a formula."""
    if value is None:
        return ""
    if not isinstance(value, str):
        return value
    if value[:1] in _DANGEROUS_PREFIX:
        return "'" + value
    return value


def select_leads(results, want, include_cool=False):
    """
    Keep the leads worth calling, best first.

    'hot'  = advertising-related infrastructure was found and the funnel leaks.
    'warm' = the firm builds an audience on social and the funnel leaks.
    'cool' = a quiet firm with a real defect, but no proof of marketing effort.
    """
    keep = [r for r in results if r["warm"] and not r["disqualified"]]
    if not include_cool:
        keep = [r for r in keep if r.get("tier") != scoring.TIER_COOL] or keep
    keep.sort(key=scoring.sort_key)
    return keep[:want]


def write_csv(rows, path):
    with open(path, "w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=CSV_FIELDS, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({key: safe_cell(row.get(key, "")) for key in CSV_FIELDS})
    return path


def _lead_state_key(row):
    """
    Produce a stable, opaque key identifying this business across runs and rankings.

    Used by the call sheet so contacted state can survive re-ordering, scoring
    changes, and filter adjustments.
    """
    website = (row.get("website") or "").strip().lower()
    if website:
        identity = f"website:{website}"
    else:
        phone_digits = "".join(c for c in (row.get("phone") or "") if c.isdigit())
        if phone_digits:
            identity = f"phone:{phone_digits}"
        else:
            name = " ".join((row.get("name") or "").strip().lower().split())
            identity = f"name:{name}"

    digest = hashlib.sha256(identity.encode("utf-8")).hexdigest()[:16]
    return f"lead-{digest}"


# ---------------------------------------------------------------------------
# The HTML call sheet
# ---------------------------------------------------------------------------

_PAGE = """<!doctype html>
<html lang="en"><head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>LeadScan call sheet</title>
<style>
  :root {{ color-scheme: light dark;
    --bg:#f6f7f9; --card:#fff; --ink:#16181d; --muted:#5c6472; --line:#e3e6ec;
    --hot:#c2410c; --hotbg:#fff2e8; --warm:#a16207; --warmbg:#fef6e0;
    --cool:#15803d; --coolbg:#f0fdf4; --brand:#0f766e;
  }}
  @media (prefers-color-scheme: dark) {{
    :root {{
      --bg:#0f1217; --card:#171b22; --ink:#e6edf3; --muted:#8b949e; --line:#30363d;
      --hot:#fb923c; --hotbg:#431407; --warm:#facc15; --warmbg:#422006;
      --cool:#4ade80; --coolbg:#052e16; --brand:#2dd4bf;
    }}
  }}
  * {{ box-sizing:border-box }}
  body {{ margin:0; background:var(--bg); color:var(--ink);
    font:14px/1.55 -apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,sans-serif;
    padding:24px 20px; }}
  header, main, footer {{ max-width:880px; margin:0 auto }}
  header {{ margin-bottom:20px }}
  h1 {{ font-size:22px; margin:0 0 4px; letter-spacing:-.01em }}
  .sub {{ color:var(--muted); font-size:13px }}
  .counts {{ display:flex; gap:8px; margin-top:12px }}
  .pill {{ padding:3px 9px; border-radius:99px; font-size:12px; font-weight:600 }}
  .pill.hot {{ background:var(--hotbg); color:var(--hot) }}
  .pill.warm {{ background:var(--warmbg); color:var(--warm) }}
  .pill.cool {{ background:var(--coolbg); color:var(--cool) }}
  main {{ display:grid; gap:12px }}
  .lead {{ background:var(--card); border:1px solid var(--line); border-radius:10px;
    padding:14px 16px; display:grid; grid-template-columns:auto 1fr auto;
    gap:14px; align-items:flex-start; transition:opacity .15s }}
  .lead.done {{ opacity:.45 }}
  .lead.done .name {{ text-decoration:line-through }}
  .tick {{ margin-top:3px; width:17px; height:17px; cursor:pointer }}
  .name {{ font-weight:700; font-size:15.5px; display:inline }}
  .rank {{ display:inline-block; color:var(--muted); font-size:12px;
    margin-right:6px; font-weight:600 }}
  .tag {{ display:inline-block; font-size:11px; font-weight:700; text-transform:uppercase;
    letter-spacing:.04em; padding:1px 6px; border-radius:4px; margin-left:6px;
    vertical-align:middle }}
  .tag.hot {{ background:var(--hotbg); color:var(--hot) }}
  .tag.warm {{ background:var(--warmbg); color:var(--warm) }}
  .tag.cool {{ background:var(--coolbg); color:var(--cool) }}
  .hook {{ margin:6px 0 6px; font-size:13.5px }}
  .meta {{ font-size:12px; color:var(--muted) }}
  .meta a {{ color:inherit; text-decoration:underline }}
  .side {{ text-align:right }}
  .tel {{ display:inline-block; background:var(--brand); color:#fff !important;
    text-decoration:none; padding:7px 12px; border-radius:6px; font-weight:600;
    font-size:13px; white-space:nowrap }}
  .score {{ color:var(--muted); font-size:11.5px; margin-top:4px }}
  .outcome {{ margin-top:6px; max-width:150px; padding:5px 7px; border:1px solid var(--line);
    border-radius:7px; background:var(--card); color:var(--ink); font:inherit; font-size:12px }}
  footer {{ margin-top:32px; color:var(--muted); font-size:12px; text-align:center }}
  @media print {{
    body {{ background:#fff; color:#000; padding:0 }}
    .lead {{ break-inside:avoid; border-color:#ccc }}
    .outcome {{ display:none }}
  }}
</style></head>
<body>
<header>
  <h1>LeadScan call sheet</h1>
  <div class="sub">{count} leads, best first. Generated {stamp}. Tick a row when the call is done.</div>
  <div class="counts">{pills}</div>
</header>
<main>
{cards}
</main>
<footer>
  These records hold business phone numbers taken from public listings. Keep the
  file private and delete it when the campaign ends.
</footer>
<script>
(function () {{
  var CALLED_KEY = 'leadscan.called.v1';
  var OUTCOME_KEY = 'leadscan.outcome.v1';
  var called = {{}};
  var outcomes = {{}};

  function isValidOutcome(val) {{
    return val === 'follow-up' || val === 'interested' || val === 'not-interested';
  }}

  try {{
    var rawCalled = localStorage.getItem(CALLED_KEY);
    if (rawCalled) {{
      var parsedCalled = JSON.parse(rawCalled);
      if (parsedCalled && typeof parsedCalled === 'object' && !Array.isArray(parsedCalled)) {{
        called = parsedCalled;
      }}
    }}
  }} catch (e) {{
    called = {{}};
  }}

  try {{
    var rawOutcomes = localStorage.getItem(OUTCOME_KEY);
    if (rawOutcomes) {{
      var parsedOutcomes = JSON.parse(rawOutcomes);
      if (parsedOutcomes && typeof parsedOutcomes === 'object' && !Array.isArray(parsedOutcomes)) {{
        outcomes = parsedOutcomes;
      }}
    }}
  }} catch (e) {{
    outcomes = {{}};
  }}

  document.querySelectorAll('.tick').forEach(function (box) {{
    var lead = box.closest('.lead');
    var key = lead ? lead.dataset.leadKey : null;

    if (key && called[key] === true) {{
      box.checked = true;
      lead.classList.add('done');
    }}

    box.addEventListener('change', function () {{
      var isDone = box.checked;
      if (lead) {{
        lead.classList.toggle('done', isDone);
      }}
      if (key) {{
        if (isDone) {{
          called[key] = true;
        }} else {{
          delete called[key];
        }}
        try {{
          localStorage.setItem(CALLED_KEY, JSON.stringify(called));
        }} catch (e) {{
          // Storage write failure is handled gracefully
        }}
      }}
    }});
  }});

  document.querySelectorAll('.outcome').forEach(function (select) {{
    var lead = select.closest('.lead');
    var key = lead ? lead.dataset.leadKey : null;

    if (key && isValidOutcome(outcomes[key])) {{
      select.value = outcomes[key];
    }}

    select.addEventListener('change', function () {{
      var val = select.value;
      if (key) {{
        if (isValidOutcome(val)) {{
          outcomes[key] = val;
        }} else {{
          delete outcomes[key];
        }}
        try {{
          localStorage.setItem(OUTCOME_KEY, JSON.stringify(outcomes));
        }} catch (e) {{
          // Storage write failure is handled gracefully
        }}
      }}
    }});
  }});
}})();
</script>
</body></html>
"""

_CARD = """  <article class="lead" data-lead-key="{lead_key}">
    <input class="tick" type="checkbox" aria-label="Mark {name} as called">
    <div>
      <div class="rank">#{rank}</div>
      <div class="name">{name}</div>
      <span class="tag {tier}">{tier}</span>{adnote}
      <p class="hook">{hook}</p>
      <div class="meta">{meta}</div>
    </div>
    <div class="side">
      {call}
      <div class="score">score {score}</div>
      <select class="outcome" aria-label="Contact outcome for {name}">
        <option value="">No outcome</option>
        <option value="follow-up">Follow up</option>
        <option value="interested">Interested</option>
        <option value="not-interested">Not interested</option>
      </select>
    </div>
  </article>"""


def write_html(rows, path, stamp=""):
    cards = []
    for index, row in enumerate(rows, 1):
        tier = row.get("tier") or "cool"
        phone = (row.get("phone") or "").strip()
        if phone:
            call = (f'<a class="tel" href="tel:{html_module.escape(_tel(phone))}">'
                    f'{html_module.escape(phone)}</a>')
        else:
            call = '<span class="tel">no phone listed</span>'

        meta = []
        if row.get("review_count") is not None:
            meta.append(f"{row['review_count']} reviews")
        if row.get("instagram_followers"):
            meta.append(f"{row['instagram_followers']:,} IG followers")
        if row.get("email"):
            label = html_module.escape(row["email"])
            note = (row.get("email_grade") or "").strip()
            if note and note != "personal":
                label += f" ({html_module.escape(note)})"
            meta.append(f'<a href="mailto:{html_module.escape(row["email"])}">'
                        f'{label}</a>')
        for label, key in (("website", "website"), ("Instagram", "instagram"),
                           ("Facebook", "facebook"), ("TikTok", "tiktok")):
            link = (row.get(key) or "").strip()
            if link:
                meta.append(f'<a href="{html_module.escape(link)}" target="_blank" '
                            f'rel="noopener noreferrer">{label}</a>')
        hours = (row.get("opening_hours") or "").strip()
        if hours:
            meta.append(f'<span title="{html_module.escape(hours)}">opening hours &#9432;</span>')
        if row.get("status") and row["status"] != "ok":
            meta.append(html_module.escape(str(row["status"])))

        chat = (row.get("whatsapp_link") or "").strip()
        if chat:
            meta.append(f'<a href="{html_module.escape(chat)}" target="_blank" '
                        f'rel="noopener noreferrer">open WhatsApp with the '
                        f'message ready</a>')
        adtags = (row.get("ad_tags") or "").strip()
        adnote = (f' <span class="score">{html_module.escape(adtags)}</span>'
                  if adtags else "")

        cards.append(_CARD.format(
            lead_key=html_module.escape(_lead_state_key(row)),
            rank=index,
            name=html_module.escape(row.get("name", "")),
            tier=html_module.escape(tier),
            adnote=adnote,
            hook=html_module.escape(row.get("hook", "")),
            meta=" &middot; ".join(meta),
            call=call,
            score=row.get("score", 0),
        ))

    counts = {}
    for row in rows:
        counts[row.get("tier") or "cool"] = counts.get(row.get("tier") or "cool", 0) + 1
    pills = "".join(
        f'<span class="pill {tier}">{counts[tier]} {tier}</span>'
        for tier in ("hot", "warm", "cool") if counts.get(tier)
    )

    page = _PAGE.format(
        count=len(rows),
        stamp=html_module.escape(stamp),
        pills=pills or '<span class="pill">no leads</span>',
        cards="\n".join(cards) if cards else '<p style="padding:20px;text-align:center;color:var(--muted)">No leads matched the filters.</p>',
    )
    with open(path, "w", encoding="utf-8") as handle:
        handle.write(page)
    return path


def _tel(phone):
    """Make a dialable tel: value. Keep a leading plus, drop everything else."""
    digits = "".join(ch for ch in phone if ch.isdigit())
    return ("+" if phone.strip().startswith("+") else "") + digits


# ---------------------------------------------------------------------------
# The spreadsheet
# ---------------------------------------------------------------------------

def write_xlsx(rows, path):
    """Write the call sheet as a real spreadsheet. Needs openpyxl."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    tier_fill = {
        "hot": PatternFill("solid", fgColor="FFE7D6"),
        "warm": PatternFill("solid", fgColor="FDF1D0"),
        "cool": PatternFill("solid", fgColor="EAF2DC"),
    }
    widths = {"Business": 30, "What to say": 78, "Phone": 18, "Website": 34,
              "Email": 28, "Advertising tags installed": 24, "Opening hours": 40,
              "Instagram": 30, "Facebook": 30,
              "TikTok": 30, "Status": 16}

    book = Workbook()
    sheet = book.active
    sheet.title = "Call sheet"

    headings = [column[0] for column in SHEET_COLUMNS]
    sheet.append(headings)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F3B52")
        cell.alignment = Alignment(vertical="center")
    sheet.freeze_panes = "A2"

    for row in rows:
        values = []
        for _, key, _is_link in SHEET_COLUMNS:
            values.append(safe_cell(row.get(key, "")))
        sheet.append(values)
        fill = tier_fill.get(row.get("tier"))
        if fill:
            sheet.cell(row=sheet.max_row, column=1).fill = fill

    # Make the links clickable.
    for index, (_heading, key, is_link) in enumerate(SHEET_COLUMNS, start=1):
        if not is_link:
            continue
        for number, row in enumerate(rows, start=2):
            link = (row.get(key) or "").strip()
            if link.startswith(("http://", "https://")):
                cell = sheet.cell(row=number, column=index)
                cell.hyperlink = link
                cell.style = "Hyperlink"

    for index, heading in enumerate(headings, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = widths.get(heading, 14)
    for number in range(2, sheet.max_row + 1):
        sheet.cell(row=number, column=5).alignment = Alignment(wrap_text=True, vertical="top")
        sheet.row_dimensions[number].height = 42

    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headings))}{sheet.max_row}"
    book.save(path)
    return path


def write_all(rows, base_path, stamp=""):
    """Write every format next to `base_path`. Give back the paths written."""
    stem, _ = os.path.splitext(base_path)
    folder = os.path.dirname(os.path.abspath(stem))
    if folder:
        os.makedirs(folder, exist_ok=True)
    written = {"csv": write_csv(rows, stem + ".csv"),
               "html": write_html(rows, stem + ".html", stamp)}
    xlsx = write_xlsx(rows, stem + ".xlsx")
    if xlsx:
        written["xlsx"] = xlsx
    return written
